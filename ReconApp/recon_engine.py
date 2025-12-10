# recon_engine.py
import pandas as pd
from pathlib import Path
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image
import openpyxl.utils
import time

#Update#
# === CONFIG / Defaults ===
TOLERANCE = 0.001

# By default mapping and plc files are expected in ./static/
BASE_DIR = Path(__file__).parent
STATIC_DIR = BASE_DIR / "static"
DEFAULT_MAPPING = STATIC_DIR / "mapping_data3.xlsx"
DEFAULT_PLC = STATIC_DIR / "PLC_data.xlsx"


# === HELPERS ===
def normalize_account(val):
    if pd.isna(val):
        return ""
    return "".join(ch for ch in str(val) if ch.isdigit())


def to_float(val):
    if pd.isna(val):
        return 0.0
    val = str(val).replace(",", ".").replace(" ", "")
    try:
        return round(float(val), 2)
    except:
        return 0.0

def hyperlink_to_frontpage(cell):
    """Convert a cell into a hyperlink pointing back to the frontpage."""
    cell.hyperlink = "#Frontpage!A1"
    cell.style = "Hyperlink"


def _normalize_code(x):
    """Return consistent code strings (e.g., 101, 101.0, ' 101 ') -> '101'."""
    if pd.isna(x):
        return None
    try:
        f = float(x)
        return str(int(f)) if f.is_integer() else str(f).rstrip("0").rstrip(".")
    except:
        s = str(x).strip()
        if s.replace(".", "", 1).isdigit():
            try:
                f = float(s)
                return str(int(f)) if f.is_integer() else s
            except:
                pass
        return s


# Accounts that must use ICP–totals logic (same as 731000 & 321000)
ICP_TOTAL_ACCOUNTS = {
    "731000",
    "321000",
    "321001",
    "731001",
    "634010",
    "234110",
}


# === MAPPING LOAD/APPLY ===
def load_mapping(mapping_path=None):
    """
    Returns: acct_to_code (dict), code_to_meta (dict), map_dir (DataFrame)
    Expects mapping_path to be an Excel workbook with sheets:
      - account_mapping (with columns: Account no., Mapping)
      - mapping_directory (with columns: code, header, sheet)
    """
    mapping_path = Path(mapping_path) if mapping_path else DEFAULT_MAPPING
    if not mapping_path.exists():
        raise FileNotFoundError(f"Mapping file not found at {mapping_path}")

    book = pd.read_excel(mapping_path, sheet_name=None)

    if "account_mapping" not in book or "mapping_directory" not in book:
        raise KeyError("mapping_data3.xlsx must include sheets 'account_mapping' and 'mapping_directory'")

    map_accounts = book["account_mapping"].rename(columns={"Mapping": "code"}).copy()
    map_dir = book["mapping_directory"].copy()

    # Normalize
    map_accounts["Account no."] = map_accounts["Account no."].astype(str).str.strip().apply(normalize_account)
    map_accounts["code"] = map_accounts["code"].apply(_normalize_code)

    map_dir["code"] = map_dir["code"].apply(_normalize_code)
    map_dir["header"] = map_dir["header"].astype(str).str.strip()
    map_dir["sheet"] = map_dir["sheet"].astype(str).str.strip()

    acct_to_code = map_accounts.set_index("Account no.")["code"].to_dict()
    code_to_meta = map_dir.set_index("code")[["header", "sheet"]].to_dict(orient="index")

    return acct_to_code, code_to_meta, map_dir


def apply_mapping(trial_balance, acct_to_code, code_to_meta, bank_code="19", bank_ranges=((390000, 399999),)):
    tb = trial_balance.copy()
    tb["No."] = tb["No."].astype(str).str.strip().apply(normalize_account)
    tb["code"] = tb["No."].map(acct_to_code)

    def _in_ranges(acc_str, ranges):
        return bool(acc_str) and acc_str.isdigit() and any(lo <= int(acc_str) <= hi for lo, hi in ranges)

    unmapped = tb["code"].isna()
    tb.loc[unmapped & tb["No."].apply(lambda s: _in_ranges(s, bank_ranges)), "code"] = bank_code

    tb["header"] = tb["code"].map(lambda c: code_to_meta.get(c, {}).get("header"))
    tb["sheet_group"] = tb["code"].map(lambda c: code_to_meta.get(c, {}).get("sheet"))
    tb["sheet_group"] = tb["sheet_group"].fillna("Unmapped").astype(str).str.strip()
    return tb


# === STYLES / BORDERS / FORMATS ===
thin = Side(border_style="thin", color="000000")
thick = Side(border_style="medium", color="000000")

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
header_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # header darker
entry_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # lighter
total_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")    # totals
plc_fill = PatternFill(start_color="9AD29F", end_color="9AD29F", fill_type="solid")

def apply_borders(ws, top, bottom, left, right):
    """
    Draw neat rectangle with thick external lines and thin internal lines
    Ensures thick border above/below column A and last Amount column as requested.
    """
    # thin grid everywhere in the block
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # thick top and bottom across the full block
    for c in range(left, right + 1):
        top_cell = ws.cell(top, c)
        bottom_cell = ws.cell(bottom, c)
        top_cell.border = Border(top=thick, left=top_cell.border.left, right=top_cell.border.right, bottom=top_cell.border.bottom)
        bottom_cell.border = Border(bottom=thick, left=bottom_cell.border.left, right=bottom_cell.border.right, top=bottom_cell.border.top)

    # thick left & right sides
    for r in range(top, bottom + 1):
        left_cell = ws.cell(r, left)
        right_cell = ws.cell(r, right)
        left_cell.border = Border(left=thick, top=left_cell.border.top, bottom=left_cell.border.bottom, right=left_cell.border.right)
        right_cell.border = Border(right=thick, top=right_cell.border.top, bottom=right_cell.border.bottom, left=right_cell.border.left)

    # corners: ensure they have both thick sides
    ws.cell(top, left).border = Border(top=thick, left=thick, right=thin, bottom=thin)
    ws.cell(top, right).border = Border(top=thick, right=thick, left=thin, bottom=thin)
    ws.cell(bottom, left).border = Border(bottom=thick, left=thick, right=thin, top=thin)
    ws.cell(bottom, right).border = Border(bottom=thick, right=thick, left=thin, top=thin)

####### NEW CODE FRESHLY BAKED 

def add_pl_balance_sheet(wb, trial_balance_df, code_to_meta):
    """
    Builds the PL & Balance sheet.
    - Mapping lines (101, 102, 1, 2, 20, ...) are numeric values.
    - Subtotals / totals are Excel formulas based on those rows.
    """

    # Insert after Trial Balance
    try:
        tb_index = wb.sheetnames.index("Trial Balance (YTD)")
    except ValueError:
        tb_index = 0
    ws = wb.create_sheet("PL & Balance", tb_index + 1)

    # === Layout definitions ===
    PL_LAYOUT = [
        ("",  "Revenue", None),
        ("101", "Revenue", "Income"),
        ("102", "Results from subsidiaries", "Investments"),
        ("103", "Results from associated companies", "Investments"),
        ("104", "Other Income", "Income"),
        ("105", "Direct Costs", "Costs"),
        ("", "Gross Profit", None),

        ("106", "Staff costs", "Costs"),
        ("107", "Other external costs", "Costs"),
        ("", "EBITDA", None),

        ("108", "Depreciation & impairment", "PPE"),
        ("", "Operating Profit", None),

        ("109", "Finance income", "Finance Income & Expenses"),
        ("110", "Finance expenses", "Finance Income & Expenses"),
        ("", "Profit before tax", None),

        ("111", "Tax", "Corp. & Deferred Tax"),
        ("", "Total profit", None),
    ]

    ASSETS_LAYOUT = [
        ("", "Non-current assets", None),
        ("1", "Goodwill", "Goodwill"),
        ("2", "Property, plant and equipment", "PPE"),
        ("3", "Lease assets", "Lease Assets"),
        ("4", "Investments in subsidiaries", "Investments"),
        ("5", "Investments in associated companies", "Investments"),
        ("6", "Other investments", "Investments"),
        ("8", "Intercompany receivables", "Intercompany"),
        ("9", "Long term trade receivables and contract assets", "Trade Receivables"),
        ("10", "Derivatives, long term", "Derivatives"),
        ("11", "Other long term receivables", "Other Receivables"),
        ("12", "Deferred tax assets", "Corp. & Deferred Tax"),
        ("", "Total non-current assets", None),

        ("", "Current assets", None),
        ("13", "Inventories (development costs and projects)", "Inventories"),
        ("14", "Derivatives, short term", "Derivatives"),
        ("15", "Trade receivables and contract assets", "Trade Receivables"),
        ("16", "Other receivables", "Other Receivables"),
        ("17", "Prepayments", "Prepayments"),
        ("19", "Cash and cash equivalents", "Cash & Cash Equivalents"),
        ("", "Total current assets", None),

        ("", "Total assets", None),
    ]


    EQUITY_LIAB_LAYOUT = [
        ("", "Equity", None),
        ("20", "Share capital", "Equity"),
        ("21", "Retained earnings and reserves", "Equity"),
        ("22", "Hybrid Capital", "Equity"),
        ("23", "Minority interests", "Equity"),
        ("TP_EQUITY", "Total profit", None),
        ("", "Total equity", None),

        ("", "Liabilities", None),
        ("24", "Bonds", "Bonds"),
        ("25", "Long-term Project Financing", "Mortage Credit"),
        ("28", "Provisions", "Provisions"),
        ("29", "Derivatives, long term", "Derivatives"),
        ("30", "Deferred tax", "Corp. & Deferred Tax"),
        ("", "Total non-current liabilities", None),

        ("", "Current liabilities", None),
        ("31", "Project Financing", "Mortage Credit"),
        ("33", "Trade Payables", "Trade Payables"),
        ("34", "Derivatives, short term", "Derivatives"),
        ("35", "Intercompany payables", "Intercompany"),
        ("36", "Corporation tax", "Corp. & Deferred Tax"),
        ("37", "Provisions", "Provisions"),
        ("38", "Deferred income", "Deferred Income"),
        ("39", "Other payables", "Accruals & Other Payables"),
        ("", "Total current liabilities", None),

        ("", "Total Liabilities", None),

        ("", "Assets = Liabilities Control", None),
    ]

    # === TB totals by mapping code ===
    code_totals = trial_balance_df.groupby("code")["Balance at Date"].sum().to_dict()

    def get_code_total(code: str) -> float:
        if not code or pd.isna(code):
            return 0.0
        return float(code_totals.get(str(code), 0.0))

    # Track where each description and mapping code ended up
    desc_row = {}   # description -> row
    code_row = {}   # mapping code -> row

    def write_block(start_row, layout):
        row = start_row
        block_top = row
        prev_group = False
    
        for code, desc, tab in layout:
            is_group = (code == "")
    
            # Gap between consecutive header/total rows
            if prev_group and is_group:
                for col in range(3, 6):
                    gap_cell = ws.cell(row, col, "")
                    gap_cell.fill = entry_fill
                row += 1
            prev_group = is_group
    
            # Column B: mapping code
            ws.cell(row, 2, code if code else "")
    
            # Column C: description
            desc_cell = ws.cell(row, 3, desc)
    
            # Column D: value (placeholder)
            val_cell = ws.cell(row, 4)
    
            # === ⭐ SPECIAL CASE: Total profit inside Equity section ⭐ ===
            if code == "TP_EQUITY":
                # Description formatting
                desc_cell.font = Font(bold=True)
                desc_cell.fill = entry_fill
    
                # Value = reference main P&L total profit
                tp_main_row = desc_row["Total profit"]  # P&L total profit row
                val_cell.value = f"=D{tp_main_row}"
                val_cell.number_format = "#,##0.00"
                val_cell.fill = entry_fill
    
                # No code / no hyperlink
                ws.cell(row, 2, "")
                tab_cell = ws.cell(row, 5, "")
                tab_cell.fill = entry_fill
    
                # Track row
                desc_row[desc] = row
    
                row += 1
                continue
            # === END SPECIAL CASE ===
    
            # === STANDARD LINE HANDLING ===
    
            # Determine row fill
            if code:
                row_fill = entry_fill
            else:
                if "Total" in desc or desc in (
                    "Gross Profit", "EBITDA", "Operating Profit", "Profit before tax"
                ):
                    row_fill = total_fill
                else:
                    row_fill = header_fill
    
            desc_cell.fill = row_fill
            val_cell.fill = row_fill
    
            # Column E: Tab cell
            tab_cell = ws.cell(row, 5)
            tab_cell.fill = row_fill
    
            if code:
                v = get_code_total(code)
                val_cell.value = v
                val_cell.number_format = "#,##0.00"
    
                if tab:
                    tab_cell.value = tab
                    if abs(v) > 0.00001:
                        safe_tab = f"'{tab}'" if not tab.isalnum() else tab
                        tab_cell.hyperlink = f"#{safe_tab}!A1"
                        tab_cell.font = Font(color="0000FF", underline="single")
            else:
                tab_cell.value = ""
    
            # Bold formatting for headers/totals
            if code == "" or "Total" in desc or desc in (
                "Gross Profit", "EBITDA", "Operating Profit", "Profit before tax"
            ):
                desc_cell.font = Font(bold=True)
    
            # Track rows
            desc_row[desc] = row
            if code:
                code_row[code] = row
    
            row += 1
    
        apply_borders(ws, block_top, row - 1, 3, 5)
        return row + 1



    # === Write the three blocks ===
    r = 2
    r = write_block(r, PL_LAYOUT)
    r = write_block(r, ASSETS_LAYOUT)
    r = write_block(r, EQUITY_LIAB_LAYOUT)

    # Helper to set a formula in the D-column of a description row
    def set_formula(desc: str, formula: str):
        r_ = desc_row.get(desc)
        if not r_:
            return
        cell = ws.cell(r_, 4)
        cell.value = formula
        cell.number_format = "#,##0.00"

    # Shortcuts for "D[row]"
    def d_ref(row: int) -> str:
        return f"D{row}"

    # === P&L formulas ===
    r_gp   = desc_row["Gross Profit"]
    r_ebit = desc_row["EBITDA"]
    r_op   = desc_row["Operating Profit"]
    r_pbt  = desc_row["Profit before tax"]
    r_tp   = desc_row["Total profit"]

    r101 = code_row["101"]
    r105 = code_row["105"]
    r106 = code_row["106"]
    r107 = code_row["107"]
    r108 = code_row["108"]
    r109 = code_row["109"]
    r110 = code_row["110"]
    r111 = code_row["111"]

    # Gross profit = SUM(101–105)
    set_formula("Gross Profit", f"=SUM({d_ref(r101)}:{d_ref(r105)})")

    # EBITDA = Gross profit + (106+107)
    set_formula("EBITDA", f"={d_ref(r_gp)}+({d_ref(r106)}+{d_ref(r107)})")

    # Operating profit = EBITDA + 108
    set_formula("Operating Profit", f"={d_ref(r_ebit)}+{d_ref(r108)}")

    # Profit before tax = Operating profit + (109+110)
    set_formula("Profit before tax", f"={d_ref(r_op)}+({d_ref(r109)}+{d_ref(r110)})")

    # Total profit = Profit before tax + 111
    set_formula("Total profit", f"={d_ref(r_pbt)}+{d_ref(r111)}")

    # === Assets formulas ===
    r_tnca = desc_row["Total non-current assets"]
    r_tca  = desc_row["Total current assets"]
    r_ta   = desc_row["Total assets"]

    r1  = code_row["1"]
    r12 = code_row["12"]
    r13 = code_row["13"]
    r19 = code_row["19"]

    # Total non-current assets = SUM(1–12)
    set_formula("Total non-current assets", f"=SUM({d_ref(r1)}:{d_ref(r12)})")

    # Total current assets = SUM(13–17,19) (contiguous codes in layout)
    set_formula("Total current assets", f"=SUM({d_ref(r13)}:{d_ref(r19)})")

    # Total assets = Total non-current + Total current
    set_formula("Total assets", f"={d_ref(r_tnca)}+{d_ref(r_tca)}")

    # === Equity & Liabilities formulas ===
    r_teq   = desc_row["Total equity"]
    r_tncl  = desc_row["Total non-current liabilities"]
    r_tcl   = desc_row["Total current liabilities"]
    r_tl    = desc_row["Total Liabilities"]

    r20 = code_row["20"]
    r23 = code_row["23"]
    r24 = code_row["24"]
    r30 = code_row["30"]
    r31 = code_row["31"]
    r39 = code_row["39"]

    # Total equity = SUM(20–23) + Total profit
    r_tp_equity = desc_row["Total profit"]  # the new one inside Equity
    set_formula("Total equity", f"=SUM({d_ref(r20)}:{d_ref(r_tp_equity)})")


    # Total non-current liabilities = SUM(24,25,28,29,30) (contiguous in layout)
    set_formula("Total non-current liabilities", f"=SUM({d_ref(r24)}:{d_ref(r30)})")

    # Total current liabilities = SUM(31,33,34,35,36,37,38,39) (contiguous in layout)
    set_formula("Total current liabilities", f"=SUM({d_ref(r31)}:{d_ref(r39)})")

    # Total Liabilities = Total equity + Total non-current + Total current
    set_formula("Total Liabilities", f"={d_ref(r_teq)}+{d_ref(r_tncl)}+{d_ref(r_tcl)}")

    # === Assets = Liabilities Control ===
    r_ctrl = desc_row["Assets = Liabilities Control"]
    set_formula("Assets = Liabilities Control", f"={d_ref(r_ta)}+{d_ref(r_tl)}")

    # Colour control row (green if ~0, red otherwise) using Python-calculated values
    # (same logic as the formulas, just evaluated once)
    def sum_codes(code_list):
        return sum(get_code_total(c) for c in code_list)

    gross_profit_val = sum_codes(["101", "102", "103", "104", "105"])
    ebitda_val = gross_profit_val + sum_codes(["106", "107"])
    op_val = ebitda_val + get_code_total("108")
    pbt_val = op_val + sum_codes(["109", "110"])
    total_profit_val = pbt_val + get_code_total("111")

    total_nca_val = sum_codes([str(c) for c in range(1, 13)])
    total_ca_val = sum_codes([str(c) for c in [13, 14, 15, 16, 17, 19]])
    total_assets_val = total_nca_val + total_ca_val

    total_equity_val = sum_codes([str(c) for c in [20, 21, 22, 23]]) + total_profit_val
    total_ncl_val = sum_codes([str(c) for c in [24, 25, 28, 29, 30]])
    total_cl_val = sum_codes([str(c) for c in [31, 33, 34, 35, 36, 37, 38, 39]])
    total_liab_val = total_equity_val + total_ncl_val + total_cl_val

    # Format the Assets = Liabilities Control line as a header
    for col in range(3, 6):
        cell = ws.cell(r_ctrl, col)
        cell.fill = header_fill
    
    # Bold description only
    ws.cell(r_ctrl, 3).font = Font(bold=True)
    # Value cell not bold
    ws.cell(r_ctrl, 4).font = Font(bold=False)


    ws.column_dimensions["B"].hidden = True

   # === UNMAPPED TOTALS BLOCK (only if unmapped exists) ===
    unmapped_df = trial_balance_df[trial_balance_df["sheet_group"] == "Unmapped"]
    unmapped_total = float(unmapped_df["Balance at Date"].sum()) if not unmapped_df.empty else 0.0
    
    if abs(unmapped_total) > 0.01:
    
        # ---- GAP ROW ----
       # Insert gap row before unmapped box (blank white row)
        gap_row = r_ctrl + 1
        for col in range(3, 6):
            gap_cell = ws.cell(gap_row, col, "")
            gap_cell.fill = PatternFill(fill_type=None)   

    
        # Box starts 1 row below the gap
        box_top = r_ctrl + 2
        row = box_top
    
        # ---- Total unmapped ----
        ws.cell(row, 3, "Total unmapped").font = Font(bold=True)
        ws.cell(row, 3).fill = header_fill
    
        unm_cell = ws.cell(row, 4, unmapped_total)
        unm_cell.number_format = "#,##0.00"
        unm_cell.fill = header_fill
    
        # ---- Diff row ----
        row += 1
        diff_label = ws.cell(row, 3, "Diff.")
        diff_label.fill = entry_fill
    
        diff_cell = ws.cell(row, 4, f"=D{r_ctrl}+D{box_top}")
        diff_cell.number_format = "#,##0.00"
        diff_cell.fill = entry_fill
    
        
        # Format "Diff." line as header-style row
        diff_label.fill = header_fill
        diff_cell.fill = header_fill
        
        # Bold label only
        diff_label.font = Font(bold=True)
        diff_cell.font = Font(bold=False)
    
        # Apply borders
        apply_borders(ws, box_top, row, 3, 4)


####### - 


# === INTERNAL ZEROING ===
def remove_internal_zeroes(df, tol=TOLERANCE):
    """
    Remove internal cancelling entries when:
      - ICP CODE is the same (or empty/NaN in both lines), and
      - GAAP Code is the same (or empty/NaN in both lines), and
      - Amounts sum to ~0 within tolerance.

    Document No. is intentionally ignored.

    Then, within each (ICP, GAAP) bucket, perform cumulative zero-block
    trimming: drop entries up to the last point where that bucket's
    cumulative sum returns to (approx.) zero.
    """
    if df.empty:
        return df

    # Sort chronologically
    df = df.sort_values("Posting Date", ascending=True).reset_index(drop=True)

    def _norm_key(x):
        """Normalize ICP/GAAP for grouping & equality: treat NaN/empty as ''."""
        if pd.isna(x):
            return ""
        return str(x).strip()

    # Normalized keys used for both pairwise and cumulative logic
    df["_icp_norm"] = df["ICP CODE"].apply(_norm_key) if "ICP CODE" in df.columns else ""
    df["_gaap_norm"] = df["GAAP Code"].apply(_norm_key) if "GAAP Code" in df.columns else ""

    # ---------- 1) Pairwise zero removal within same ICP/GAAP ----------
    keep = [True] * len(df)
    for i in range(len(df)):
        if not keep[i]:
            continue
        for j in range(i + 1, len(df)):
            if not keep[j]:
                continue

            same_icp = (df.loc[i, "_icp_norm"] == df.loc[j, "_icp_norm"])
            same_gaap = (df.loc[i, "_gaap_norm"] == df.loc[j, "_gaap_norm"])

            if same_icp and same_gaap and abs(df.loc[i, "Amount (LCY)"] + df.loc[j, "Amount (LCY)"]) < tol:
                # Exact opposite pair within same ICP/GAAP bucket -> drop both
                keep[i] = keep[j] = False
                break

    df = df[keep].copy()

    # ---------- 2) Cumulative zero trimming per (ICP, GAAP) bucket ----------
    def _trim_group(g):
        if g.empty:
            return g
        g = g.copy()
        g["cum"] = g["Amount (LCY)"].cumsum().round(2)
        zero_idx = g.index[g["cum"].abs() < tol].tolist()
        if zero_idx:
            last_zero = zero_idx[-1]
            g = g.loc[g.index > last_zero]
        g = g.drop(columns=["cum"])
        return g

    df = df.groupby(["_icp_norm", "_gaap_norm"], group_keys=False).apply(_trim_group)

    # Clean up helper columns and reindex
    df = df.drop(columns=["_icp_norm", "_gaap_norm"], errors="ignore").reset_index(drop=True)

    return df



# === WORKBOOK BUILDING #
def build_workbook(trial_balance_df, entries_df, map_dir, acct_to_code, code_to_meta, ICP, tolerance=TOLERANCE):
    """
    Returns: openpyxl.Workbook object, sheet_status dict, account_anchor dict, mismatch_accounts list
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # === TRIAL BALANCE TAB ===
    ws_tb = wb.create_sheet("Trial Balance (YTD)", 0)
    for c_idx, col in enumerate(trial_balance_df.columns, 1):
        ws_tb.cell(1, c_idx, col).font = Font(bold=True)
    for r_idx, row in enumerate(trial_balance_df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            ws_tb.cell(r_idx, c_idx, val)

    # Autofit for trial balance
    for col in ws_tb.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws_tb.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max_len + 2

    # PL & Balance sheet
    add_pl_balance_sheet(wb, trial_balance_df, code_to_meta)

    # === ACCOUNT OVERVIEW (mapped groups) ===
    sheet_order = list(map_dir["sheet"].unique()) + ["Unmapped"]
    cols_needed = ["Posting Date", "Description", "Document No.", "ICP CODE", "GAAP Code", "Amount (LCY)"]
    cols_present = [c for c in cols_needed if c in entries_df.columns]

    sheet_status = {}
    account_anchor = {}
    mismatch_accounts = []

    # iterate mapping order
    for sheet_name in sheet_order:
        subset = trial_balance_df[trial_balance_df["sheet_group"] == sheet_name]
        if subset.empty:
            continue

        ws = wb.create_sheet(sheet_name[:31])
        row_cursor = 3
        sheet_mismatch = False
        account_count = 0

        for _, tb in subset.sort_values("No.").iterrows():
            acc_no = str(tb["No."])
            acc_name = tb.get("Name", "")
            tb_bal = tb.get("Balance at Date", 0.0)

            acc_df = entries_df[entries_df["G/L Account No."] == acc_no].copy()
            if acc_df.empty:
                continue

            account_count += 1

            # Year trimming: drop prior years that net to 0 (keep latest year always)
            acc_df = acc_df.sort_values("Posting Date", ascending=True).reset_index(drop=True)
            acc_df["Year"] = pd.to_datetime(acc_df["Posting Date"], errors="coerce").dt.year.fillna(0).astype(int)
            if acc_df["Year"].nunique() > 1:
                years = sorted(acc_df["Year"].unique())
                for y in years[:-1]:
                    y_sum = round(acc_df.loc[acc_df["Year"] == y, "Amount (LCY)"].sum(), 2)
                    if abs(y_sum) < tolerance:
                        acc_df = acc_df[acc_df["Year"] != y]

            if acc_df.empty:
                continue

            # remove internal zeroes (only when Document No., ICP and GAAP match)
            acc_df = remove_internal_zeroes(acc_df)
            if acc_df.empty:
                continue

            # === Special accounts using ICP totals logic ===
            if acc_no in ICP_TOTAL_ACCOUNTS:
                tmp = acc_df.copy()

                # Create a grouping key that also captures lines WITHOUT ICP
                if "ICP CODE" in tmp.columns:
                    tmp["_icp_group"] = tmp["ICP CODE"].astype(str)
                    tmp["_icp_group"] = tmp["_icp_group"].replace("nan", "").fillna("")
                    tmp.loc[tmp["_icp_group"].str.strip() == "", "_icp_group"] = "No ICP"
                else:
                    # If column doesn't exist at all, everything is "No ICP"
                    tmp["_icp_group"] = "No ICP"

                grouped = tmp.groupby("_icp_group", as_index=False)["Amount (LCY)"].sum()
                grouped.rename(columns={"_icp_group": "ICP CODE"}, inplace=True)

                # For display: Description = ICP code (or "No ICP"), GAAP/Doc blank
                grouped["Description"] = grouped["ICP CODE"]
                grouped["Document No."] = ""
                grouped["GAAP Code"] = ""

                acc_view = grouped[["Description", "Document No.", "ICP CODE", "GAAP Code", "Amount (LCY)"]].copy()
                net_sum = round(acc_view["Amount (LCY)"].sum(), 2)

                # Header row
                header_cell = ws.cell(row=row_cursor, column=1, value=f"{acc_no} - {acc_name}")
                account_anchor[acc_no] = (ws.title, row_cursor)
                hyperlink_to_frontpage(header_cell)
                header_cell.fill = green_fill if abs(net_sum - tb_bal) <= tolerance else red_fill

                if abs(net_sum - tb_bal) > tolerance:
                    sheet_mismatch = True
                    mismatch_accounts.append({
                        "No": acc_no,
                        "Name": acc_name,
                        "tb_balance": tb_bal,
                        "entries_sum": net_sum,
                        "difference": round(net_sum - tb_bal, 2),
                    })

                row_cursor += 1
                block_start = row_cursor

                # Column headers
                cols = ["Description", "Document No.", "ICP CODE", "GAAP Code", "Amount (LCY)"]
                for c_idx, col in enumerate(cols, 1):
                    ws.cell(row=row_cursor, column=c_idx, value=col).font = Font(bold=True)
                    ws.cell(row=row_cursor, column=c_idx).fill = header_fill
                row_cursor += 1

                # Rows
                for _, r in acc_view.iterrows():
                    for c_idx, col in enumerate(cols, 1):
                        cell = ws.cell(row=row_cursor, column=c_idx, value=r.get(col, ""))
                        cell.fill = entry_fill
                    row_cursor += 1

                # Total row
                ws.cell(row=row_cursor, column=4, value="Account Total").font = Font(bold=True)
                vcell = ws.cell(row=row_cursor, column=5, value=net_sum)
                vcell.font = Font(bold=True)
                for c in range(1, 6):
                    ws.cell(row=row_cursor, column=c).fill = total_fill

                apply_borders(ws, block_start, row_cursor, 1, 5)
                row_cursor += 3
                continue

            # Special: accounts 311000 and 721000 → show ONLY one total line (no ICP grouping)
            if acc_no in ["311000", "721000"]:
                net_sum = round(acc_df["Amount (LCY)"].sum(), 2)

                header_cell = ws.cell(row=row_cursor, column=1, value=f"{acc_no} - {acc_name}")
                account_anchor[acc_no] = (ws.title, row_cursor)
                hyperlink_to_frontpage(header_cell)
                header_cell.fill = green_fill if abs(net_sum - tb_bal) <= tolerance else red_fill
                if abs(net_sum - tb_bal) > tolerance:
                    sheet_mismatch = True
                    mismatch_accounts.append({
                        "No": acc_no,
                        "Name": acc_name,
                        "tb_balance": tb_bal,
                        "entries_sum": net_sum,
                        "difference": round(net_sum - tb_bal, 2),
                    })

                row_cursor += 1
                block_start = row_cursor

                # First row: labels (Note header + Account Total label)
                ws.cell(row=row_cursor, column=1, value="Note").font = Font(bold=True)
                ws.cell(row=row_cursor, column=6, value="Account Total").font = Font(bold=True)
                for c in range(1, 7):
                    ws.cell(row=row_cursor, column=c).fill = total_fill
                row_cursor += 1

                # Second row: content - (See documentation) + total
                ws.cell(row=row_cursor, column=1, value="(See documentation)")
                vcell = ws.cell(row=row_cursor, column=6, value=net_sum)
                vcell.number_format = "#,##0.00"
                for c in range(1, 7):
                    ws.cell(row=row_cursor, column=c).fill = entry_fill

                apply_borders(ws, block_start, row_cursor, 1, 6)
                row_cursor += 3
                continue

            # Special: bank/cash accounts 390000–399999 -> totals only with note (See documentation)
            if acc_no.isdigit() and 390000 <= int(acc_no) <= 399999:
                net_sum = round(acc_df["Amount (LCY)"].sum(), 2)

                header_cell = ws.cell(row=row_cursor, column=1, value=f"{acc_no} - {acc_name}")
                account_anchor[acc_no] = (ws.title, row_cursor)
                hyperlink_to_frontpage(header_cell)
                header_cell.fill = green_fill if abs(net_sum - tb_bal) <= tolerance else red_fill
                if abs(net_sum - tb_bal) > tolerance:
                    sheet_mismatch = True
                    mismatch_accounts.append({
                        "No": acc_no,
                        "Name": acc_name,
                        "tb_balance": tb_bal,
                        "entries_sum": net_sum,
                        "difference": round(net_sum - tb_bal, 2),
                    })

                row_cursor += 1
                block_start = row_cursor

                # First row: labels (Note header + Account Total label)
                ws.cell(row=row_cursor, column=1, value="Note").font = Font(bold=True)
                ws.cell(row=row_cursor, column=6, value="Account Total").font = Font(bold=True)
                for c in range(1, 7):
                    ws.cell(row=row_cursor, column=c).fill = total_fill
                row_cursor += 1

                # Second row: content - (See documentation) + total
                ws.cell(row=row_cursor, column=1, value="(See documentation)")
                vcell = ws.cell(row=row_cursor, column=6, value=net_sum)
                vcell.number_format = "#,##0.00"
                for c in range(1, 7):
                    ws.cell(row=row_cursor, column=c).fill = entry_fill

                apply_borders(ws, block_start, row_cursor, 1, 6)
                row_cursor += 3
                continue

            # Normal accounts: show full list newest -> oldest
            acc_df = acc_df.sort_values("Posting Date", ascending=False)
            net_sum = round(acc_df["Amount (LCY)"].sum(), 2)

            header_cell = ws.cell(row=row_cursor, column=1, value=f"{acc_no} - {acc_name}")
            account_anchor[acc_no] = (ws.title, row_cursor)
            hyperlink_to_frontpage(header_cell)
            header_cell.fill = green_fill if abs(net_sum - tb_bal) <= tolerance else red_fill
            if abs(net_sum - tb_bal) > tolerance:
                sheet_mismatch = True
                mismatch_accounts.append({
                    "No": acc_no,
                    "Name": acc_name,
                    "tb_balance": tb_bal,
                    "entries_sum": net_sum,
                    "difference": round(net_sum - tb_bal, 2),
                })

            row_cursor += 1
            block_start = row_cursor

            # Column headers
            for c_idx, col in enumerate(cols_present, 1):
                ws.cell(row=row_cursor, column=c_idx, value=col).font = Font(bold=True)
                ws.cell(row=row_cursor, column=c_idx).fill = header_fill
            row_cursor += 1

            # Rows: entries
            for _, e in acc_df.iterrows():
                for c_idx, col in enumerate(cols_present, 1):
                    val = e.get(col, "")
                    cell = ws.cell(row=row_cursor, column=c_idx, value=val)
                    cell.fill = entry_fill
                row_cursor += 1

            # Totals row
            ws.cell(row=row_cursor, column=len(cols_present) - 1, value="Account Total").font = Font(bold=True)
            total_cell = ws.cell(row=row_cursor, column=len(cols_present), value=net_sum)
            total_cell.font = Font(bold=True)
            for c in range(1, len(cols_present) + 1):
                ws.cell(row=row_cursor, column=c).fill = total_fill

            apply_borders(ws, block_start, row_cursor, 1, len(cols_present))
            row_cursor += 3

        sheet_status[sheet_name] = {"mismatches": int(sheet_mismatch), "accounts": account_count}

    return wb, sheet_status, account_anchor, mismatch_accounts

# === FINALIZE: front page, formatting, save to bytes ===
def finalize_workbook_to_bytes(
    wb,
    sheet_status,
    account_anchor,
    trial_balance_df,
    entries_df,
    ICP,
    plc_path=None,
    tolerance=TOLERANCE,
    mismatch_accounts=None,
    quarter=None,
):
    """
    Adds front page, number/date formatting, hides gridlines, autofit, and returns bytes buffer.
    """
    if mismatch_accounts is None:
        mismatch_accounts = []

    from warnings import filterwarnings
    filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")

    # === FRONT PAGE SHEET ===
    ws_front = wb.create_sheet("Frontpage", 0)
    ws_front["A1"] = "EE Reconciliation Overview"
    ws_front["A1"].font = Font(size=16, bold=True)

    # === PLC CARD (ICP, Company, Accountant, Controller, Quarter) ===
    plc_norm = None
    plc_path = Path(plc_path) if plc_path else DEFAULT_PLC
    try:
        if plc_path.exists():
            plc_df = pd.read_excel(plc_path, engine="openpyxl")
            plc_norm = plc_df.copy()
            plc_norm.columns = [c.strip() for c in plc_norm.columns]
            plc_norm["ICP code_norm"] = plc_norm["ICP code"].astype(str).str.strip().str.upper()
            plc_norm["Company name"] = plc_norm["Company name"].astype(str).str.strip()
            plc_norm["Accountant"] = plc_norm["Accountant"].astype(str).str.strip()
            plc_norm["Controller"] = plc_norm["Controller"].astype(str).str.strip()
    except Exception:
        plc_norm = None

    selected_icps = [ICP]
    row_ptr = 3

    for icp in selected_icps:
        icp_key = str(icp).strip().upper()
        row = plc_norm.loc[plc_norm["ICP code_norm"] == icp_key] if plc_norm is not None else pd.DataFrame()

        # Labels
        ws_front.cell(row_ptr, 1, "ICP code").font = Font(bold=True)
        ws_front.cell(row_ptr + 1, 1, "Company name").font = Font(bold=True)
        ws_front.cell(row_ptr + 2, 1, "Accountant").font = Font(bold=True)
        ws_front.cell(row_ptr + 3, 1, "Controller").font = Font(bold=True)
        ws_front.cell(row_ptr + 4, 1, "Current Quarter").font = Font(bold=True)

        # Values
        ws_front.cell(row_ptr, 2, icp_key)
        if not row.empty:
            ws_front.cell(row_ptr + 1, 2, row.iloc[0]["Company name"])
            ws_front.cell(row_ptr + 2, 2, row.iloc[0]["Accountant"])
            ws_front.cell(row_ptr + 3, 2, row.iloc[0]["Controller"])
        else:
            ws_front.cell(row_ptr + 1, 2, "Not found in PLC.xlsx")
            ws_front.cell(row_ptr + 2, 2, "—")
            ws_front.cell(row_ptr + 3, 2, "—")

        if quarter:
            ws_front.cell(row_ptr + 4, 2, str(quarter))

        # Colour PLC block
        for r in range(row_ptr, row_ptr + 5):  # rows row_ptr..row_ptr+4
            for c in range(1, 3):              # cols A:B
                ws_front.cell(r, c).fill = plc_fill

        # Border around PLC + Quarter
        apply_borders(ws_front, top=row_ptr, bottom=row_ptr + 4, left=1, right=2)

        row_ptr += 7  # blank rows after PLC card

    # === AUTOMATICALLY GENERATED COMMENTS BOX ===
    # Quick checks
    comments = []

    mask_200_399 = (
        trial_balance_df["No."].astype(str).str.isdigit() &
        trial_balance_df["No."].astype(int).between(200000, 399999)
    )
    negatives = trial_balance_df.loc[
        mask_200_399 & (trial_balance_df["Balance at Date"] < 0),
        ["No.", "Name", "Balance at Date"]
    ]

    mask_400_plus = (
        trial_balance_df["No."].astype(str).str.isdigit() &
        (trial_balance_df["No."].astype(int) >= 400000)
    )
    positives = trial_balance_df.loc[
        mask_400_plus & (trial_balance_df["Balance at Date"] > 0),
        ["No.", "Name", "Balance at Date"]
    ]

    # Unmapped accounts = rows with no mapping code
    if "code" in trial_balance_df.columns:
        mask_unmapped = trial_balance_df["code"].isna()
        unmapped_accounts = trial_balance_df.loc[
            mask_unmapped, ["No.", "Name", "Balance at Date"]
        ]
    elif "sheet_group" in trial_balance_df.columns:
        # Fallback: use sheet_group if code is not present
        mask_unmapped = trial_balance_df["sheet_group"].astype(str).str.strip().eq("Unmapped")
        unmapped_accounts = trial_balance_df.loc[
            mask_unmapped, ["No.", "Name", "Balance at Date"]
        ]
    else:
        unmapped_accounts = pd.DataFrame(columns=["No.", "Name", "Balance at Date"])



    if not negatives.empty:
        comments.append(f"{len(negatives)} account(s) in the 200000–399999 range have negative balances.")
    if not positives.empty:
        comments.append(f"{len(positives)} account(s) in the 400000+ range have positive balances.")
    if mismatch_accounts:
        comments.append(f"{len(mismatch_accounts)} account(s) have entry totals that do not match the trial balance.")
    if not unmapped_accounts.empty:
        comments.append(f"{len(unmapped_accounts)} account(s) are unmapped and need mapping.")

    mismatched_sheets = sum(v["mismatches"] for v in sheet_status.values()) if sheet_status else 0
    if mismatched_sheets > 0:
        comments.append(
            f"{mismatched_sheets} sheet(s) contain out-of-balance accounts exceeding tolerance {tolerance}."
        )

    if not comments:
        comments.append("All sheets appear balanced within tolerance limits.")

    comments_top = row_ptr
    header_cell = ws_front.cell(row_ptr, 1, "Comments:")
    header_cell.font = Font(bold=True, underline="single")
    #Only colour column A
    ws_front.cell(row_ptr, 1).fill = header_fill
    row_ptr += 1
    
    start_row = row_ptr
    for i, comment in enumerate(comments, start=start_row):
        ws_front.cell(i, 1).fill = entry_fill
        ws_front.cell(i, 1, f"• {comment}")
    comments_bottom = start_row + len(comments) - 1
    
    # Border only around column A
    apply_borders(ws_front, top=comments_top, bottom=comments_bottom, left=1, right=1)

    row_ptr = comments_bottom + 2
    
    # --- ✅ INSERT EE LOGO HERE ✅ ---
    logo_path = STATIC_DIR / "logo.png"
    if logo_path.exists():
        try:
            logo = Image(str(logo_path))
            logo.width = 130   # adjust if needed
            logo.height = 130
            ws_front.add_image(logo, "D2")  # position on the sheet
        except Exception:
            pass  # Don't break the app if the logo fails
    # --- ✅ END LOGO BLOCK ✅ ---


    
    # === Helper for hyperlinks ===
    def set_hyperlink(cell, acc_no):
        acc = str(acc_no)
        if acc in account_anchor:
            sheet_name, anchor_row = account_anchor[acc]
            sheet_ref = f"'{sheet_name}'" if not sheet_name.isalnum() else sheet_name
            cell.hyperlink = f"#{sheet_ref}!A{anchor_row}"
            cell.style = "Hyperlink"
    
   
    # === 1) ACCOUNTS OUT OF BALANCE (RED BLOCK) ===
    if mismatch_accounts:
        block_top = row_ptr
        title_cell = ws_front.cell(row_ptr, 1, "Accounts out of balance (TB vs entries):")
        title_cell.font = Font(bold=True)
        for c in range(1, 6):
            ws_front.cell(row_ptr, c).fill = red_fill
        row_ptr += 1

        headers = ["Account", "Name", "TB balance", "Entries sum", "Difference"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_front.cell(row_ptr, col_idx, h)
            cell.font = Font(bold=True)
            cell.fill = red_fill
        row_ptr += 1

        for m in mismatch_accounts:
            acc = str(m["No"])
            c = ws_front.cell(row_ptr, 1, acc)
            set_hyperlink(c, acc)
            ws_front.cell(row_ptr, 2, m.get("Name", ""))

            tb_cell = ws_front.cell(row_ptr, 3, m.get("tb_balance", 0.0))
            ent_cell = ws_front.cell(row_ptr, 4, m.get("entries_sum", 0.0))
            diff_cell = ws_front.cell(row_ptr, 5, m.get("difference", 0.0))

            tb_cell.number_format = "#,##0.00"
            ent_cell.number_format = "#,##0.00"
            diff_cell.number_format = "#,##0.00"

            for c in range(1, 6):
                ws_front.cell(row_ptr, c).fill = red_fill

            row_ptr += 1

        block_bottom = row_ptr - 1
        apply_borders(ws_front, top=block_top, bottom=block_bottom, left=1, right=5)
        row_ptr += 1  # spacing after block

  
    # === 1b) UNMAPPED ACCOUNTS (RED BLOCK) ===
    if not unmapped_accounts.empty:
        block_top = row_ptr
    
        # Header line
        title_cell = ws_front.cell(row_ptr, 1, "Unmapped accounts (no mapping code):")
        title_cell.font = Font(bold=True)
        for c in range(1, 4):
            ws_front.cell(row_ptr, c).fill = red_fill
        row_ptr += 1
    
        # Column headers
        headers = ["Account", "Name", "TB balance"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_front.cell(row_ptr, col_idx, h)
            cell.font = Font(bold=True)
            cell.fill = red_fill
        row_ptr += 1
    
        # Data rows
        for _, r in unmapped_accounts.iterrows():
            acc = str(r["No."])
    
            # Account number + hyperlink to the Unmapped tab
            acc_cell = ws_front.cell(row_ptr, 1, acc)
            acc_cell.hyperlink = "#'Unmapped'!A1"
            acc_cell.font = Font(color="0000FF", underline="single")
            acc_cell.style = "Hyperlink"
    
            # Name + Balance
            ws_front.cell(row_ptr, 2, r.get("Name", ""))
            val_cell = ws_front.cell(row_ptr, 3, r["Balance at Date"])
            val_cell.number_format = "#,##0.00"
    
            # Row formatting
            for c in range(1, 4):
                ws_front.cell(row_ptr, c).fill = red_fill
    
            row_ptr += 1
    
        # Apply borders
        block_bottom = row_ptr - 1
        apply_borders(ws_front, top=block_top, bottom=block_bottom, left=1, right=3)
    
        row_ptr += 1  # gap after block



    # === 2) NEGATIVE BALANCES ===
    if not negatives.empty:
        neg_top = row_ptr
        title_cell = ws_front.cell(row_ptr, 1, "Negative balances (200000–399999):")
        title_cell.font = Font(bold=True)
        for c in range(1, 4):
            ws_front.cell(row_ptr, c).fill = header_fill
        row_ptr += 1

        for _, r in negatives.iterrows():
            acc = str(r["No."])
            c = ws_front.cell(row_ptr, 1, acc)
            set_hyperlink(c, acc)
            ws_front.cell(row_ptr, 2, r.get("Name", ""))
            val_cell = ws_front.cell(row_ptr, 3, r["Balance at Date"])
            val_cell.number_format = "#,##0.00"
            for col in range(1, 4):
                ws_front.cell(row_ptr, col).fill = entry_fill
            row_ptr += 1

        neg_bottom = row_ptr - 1
        apply_borders(ws_front, top=neg_top, bottom=neg_bottom, left=1, right=3)
        row_ptr += 1

    # === 3) POSITIVE BALANCES ===
    if not positives.empty:
        pos_top = row_ptr
        title_cell = ws_front.cell(row_ptr, 1, "Positive balances (400000+):")
        title_cell.font = Font(bold=True)
        for c in range(1, 4):
            ws_front.cell(row_ptr, c).fill = header_fill
        row_ptr += 1

        for _, r in positives.iterrows():
            acc = str(r["No."])
            c = ws_front.cell(row_ptr, 1, acc)
            set_hyperlink(c, acc)
            ws_front.cell(row_ptr, 2, r.get("Name", ""))
            val_cell = ws_front.cell(row_ptr, 3, r["Balance at Date"])
            val_cell.number_format = "#,##0.00"
            for col in range(1, 4):
                ws_front.cell(row_ptr, col).fill = entry_fill
            row_ptr += 1

        pos_bottom = row_ptr - 1
        apply_borders(ws_front, top=pos_top, bottom=pos_bottom, left=1, right=3)

    # === 4) DOCUMENTATION CHECKLIST ===
    doc_rules = [
        ([(142110, 142120), (142210, 142220), (143110, 143120), (144110, 144120)],
         "Add documentation for Depreciation fixed assets"),
        ([(234110, 234120)], "Add documentation for Long-term receivables"),
        ([(311000, 311020)], "Add documentation for Trade receivables"),
        ([(321000, 321100)], "Add documentation for Amounts owed by affiliate companies"),
        ([(391010, 391070), (393001, 393998)], "Add documentation for Bank account"),
        ([(634010, 634011)], "Add documentation for Other I/C Loans"),
        ([(721000, 721001)], "Add documentation for Trade payables"),
        ([(731000, 731100)], "Add documentation for Amounts owed to affiliated companies"),
    ]

    def _in_any_range(acc_int, ranges):
        return any(lo <= acc_int <= hi for lo, hi in ranges)

    doc_items = []
    for _, r in trial_balance_df.iterrows():
        acc_str = str(r["No."])
        if not acc_str.isdigit():
            continue
        acc_int = int(acc_str)
        bal = r.get("Balance at Date", 0.0) or 0.0
        if abs(bal) <= tolerance:
            continue
        for ranges, message in doc_rules:
            if _in_any_range(acc_int, ranges):
                doc_items.append({"No": acc_str, "Name": r.get("Name", ""), "Message": message})
                break

    if doc_items:
        row_ptr += 2
        ws_front.cell(row_ptr, 1, "Documentation checklist:").font = Font(bold=True, underline="single")
        row_ptr += 1

        headers = ["Account", "Name", "Comment", "Status"]
        doc_top = row_ptr
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_front.cell(row_ptr, col_idx, h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        row_ptr += 1

        dv = DataValidation(type="list", formula1='"Done"', allow_blank=True)
        ws_front.add_data_validation(dv)

        for item in doc_items:
            acc = item["No"]
            name = item["Name"]
            msg = item["Message"]

            acc_cell = ws_front.cell(row_ptr, 1, acc)
            set_hyperlink(acc_cell, acc)
            ws_front.cell(row_ptr, 2, name)
            ws_front.cell(row_ptr, 3, msg)

            status_cell = ws_front.cell(row_ptr, 4)
            dv.add(status_cell)

            for c in range(1, 5):
                ws_front.cell(row_ptr, c).fill = entry_fill

            formula = f'$D{row_ptr}="Done"'
            rule = FormulaRule(formula=[formula], fill=green_fill)
            ws_front.conditional_formatting.add(f"A{row_ptr}:D{row_ptr}", rule)

            row_ptr += 1

        doc_bottom = row_ptr - 1
        apply_borders(ws_front, top=doc_top, bottom=doc_bottom, left=1, right=4)

    # === FOOTER METADATA ===
    row_ptr += 2
    ws_front.cell(row_ptr, 1, f"Generated on: {pd.Timestamp.now():%Y-%m-%d %H:%M}")
    ws_front.cell(row_ptr + 1, 1, f"Tolerance used: {tolerance}")
    ws_front.cell(row_ptr + 2, 1, f"Source files expected in: {STATIC_DIR}")

    # === GLOBAL FORMATTING (all sheets) ===
    amount_fmt = "#,##0.00"
    date_fmt = "yyyy-mm-dd"

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, (pd.Timestamp,)) or (
                    hasattr(cell.value, "year") and not isinstance(cell.value, (int, float, str))
                ):
                    try:
                        pd.to_datetime(cell.value)
                        cell.number_format = date_fmt
                    except Exception:
                        pass
                if isinstance(cell.value, (int, float)):
                    cell.number_format = amount_fmt

    # Auto-fit columns
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max_len + 2

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio



# === PUBLIC: generate_reconciliation_file ===
def generate_reconciliation_file(
    trial_balance_file,
    entries_file,
    icp_code,
    mapping_path=None,
    plc_path=None,
    tolerance=TOLERANCE,
    quarter=None,
):
    """
    Main entrypoint for Streamlit app.
    Inputs trial_balance_file and entries_file may be:
      - file-like objects (UploadedFile)
      - pathlib.Path or str paths

    Returns:
      - BytesIO (seeked to 0) containing the generated workbook .xlsx
    """
    # Load mapping from static (app-internal) unless explicit path provided
    mapping_path = Path(mapping_path) if mapping_path else DEFAULT_MAPPING
    plc_path = Path(plc_path) if plc_path else DEFAULT_PLC

    # Read user inputs (pandas handles file-like objects)
    trial_balance = pd.read_excel(trial_balance_file)
    entries = pd.read_excel(entries_file)

    # --- 🔧 Normalize column names early ---
    entries.columns = [str(c).strip() for c in entries.columns]

    # Standardize Amount column: "Amount" or "Amount (LCY)" → "Amount (LCY)"
    entries.rename(
        columns=lambda c: "Amount (LCY)"
        if str(c).strip().lower() in ["amount", "amount (lcy)"]
        else c,
        inplace=True,
    )

    # Standardize ICP column: "ICP CODE" or "ICP Code" → "ICP CODE"
    entries.rename(
        columns=lambda c: "ICP CODE"
        if str(c).strip().lower() == "icp code"
        else c,
        inplace=True,
    )

    # Standardize GAAP column: any case variant → "GAAP Code"
    entries.rename(
        columns=lambda c: "GAAP Code"
        if str(c).strip().lower() == "gaap code"
        else c,
        inplace=True,
    )

    # --- ✅ Required columns AFTER normalization ---
    missing = []
    required_cols = ["G/L Account No.", "Posting Date", "Amount (LCY)", "ICP CODE", "GAAP Code"]
    for col in required_cols:
        if col not in entries.columns:
            missing.append(col)

    if missing:
        raise ValueError(
            "The uploaded All Entries file is missing required column(s): "
            + ", ".join(missing)
            + "\n\nPlease upload a correct All Entries file and try again."
        )

    # Load mapping tables
    acct_to_code, code_to_meta, map_dir = load_mapping(mapping_path)

    # Apply mapping to trial balance
    trial_balance = apply_mapping(trial_balance, acct_to_code, code_to_meta)

    # Cast types & normalize
    trial_balance["Balance at Date"] = trial_balance["Balance at Date"].apply(to_float)
    entries["G/L Account No."] = entries["G/L Account No."].apply(normalize_account)
    entries["Amount (LCY)"] = entries["Amount (LCY)"].apply(to_float)

    # Remove timestamps: keep date only (pandas -> datetime.date)
    entries["Posting Date"] = pd.to_datetime(entries["Posting Date"], errors="coerce").dt.date

    # Build workbook
    wb, sheet_status, account_anchor, mismatch_accounts = build_workbook(
        trial_balance,
        entries,
        map_dir,
        acct_to_code,
        code_to_meta,
        icp_code,
        tolerance=tolerance,
    )

    # Finalize & get bytes
    bio = finalize_workbook_to_bytes(
        wb,
        sheet_status,
        account_anchor,
        trial_balance,
        entries,
        icp_code,
        plc_path=plc_path,
        tolerance=tolerance,
        mismatch_accounts=mismatch_accounts,
        quarter=quarter,
    )

    return bio
